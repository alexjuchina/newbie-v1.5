from fastapi import FastAPI, Request, File, UploadFile, Form
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from starlette.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from typing import Optional, Union
from concurrent.futures import ThreadPoolExecutor
from urllib.parse import urlparse
import queue
import os
import json
import requests
import time
import base64
import uuid
import tempfile
import shutil
from PIL import Image
from io import BytesIO, StringIO
import os.path
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font
import csv
from datetime import datetime

# 导入配置文件
from config import VIDEO_GENERATION_CONFIG

# Import existing config defaults
from config import EP_CONFIG, SYSTEM_PROMPT, NETWORK_QA_CONFIG, PICTURE_GENERATION_CONFIG, VIDEO_GENERATION_CONFIG, MODEL_EP_OPTIONS, FULL_VERSION, VERSION, APP_NAME, JIMENG_AI_CONFIG

app = FastAPI(title="Doubao API Newbie")

# Allow local usage and simple cross-origin during development
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

templates = Jinja2Templates(directory="templates")
app.mount("/static", StaticFiles(directory="static"), name="static")

# Image upload management
UPLOADED_IMAGES = {}  # Store uploaded images temporarily
UPLOAD_DIR = tempfile.mkdtemp(prefix="uploaded_images_")
os.makedirs(UPLOAD_DIR, exist_ok=True)

def cleanup_uploaded_images():
    """Clean up uploaded images periodically"""
    current_time = time.time()
    expired_ids = []
    for image_id, image_info in UPLOADED_IMAGES.items():
        if current_time - image_info['upload_time'] > 3600:  # 1 hour expiration
            expired_ids.append(image_id)
            if os.path.exists(image_info['file_path']):
                os.remove(image_info['file_path'])
    
    for image_id in expired_ids:
        del UPLOADED_IMAGES[image_id]


def generate_unique_filename(prefix: str, extension: str = "jpg", directory: str = None) -> str:
    """生成唯一的文件名，确保不与现有文件重复
    
    Args:
        prefix: 文件名前缀
        extension: 文件扩展名
        directory: 目标目录
    
    Returns:
        唯一的文件名
    """
    # 使用时间戳和随机数生成唯一文件名
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    random_str = str(uuid.uuid4())[:8]  # 使用UUID的前8位作为随机字符串
    
    # 构建基本文件名
    base_filename = f"{prefix}_{timestamp}_{random_str}.{extension}"
    
    # 如果指定了目录，确保目录存在并返回完整路径
    if directory:
        os.makedirs(directory, exist_ok=True)
        full_path = os.path.join(directory, base_filename)
        
        # 双重检查确保文件不存在（极端情况下可能出现冲突）
        counter = 1
        while os.path.exists(full_path):
            unique_filename = f"{prefix}_{timestamp}_{random_str}_{counter}.{extension}"
            full_path = os.path.join(directory, unique_filename)
            counter += 1
            
        return full_path
    
    return base_filename


def download_image(url: str, save_dir: str = None, filename_prefix: str = "generated_image") -> str:
    """Download image to local storage with unique filename generation and resumable download support
    
    Args:
        url: Image URL
        save_dir: Save directory, uses temp directory if None
        filename_prefix: Filename prefix
    
    Returns:
        Path to the saved image
    """
    # 确定保存目录
    if not save_dir:
        save_dir = tempfile.gettempdir()
    
    # 确保目录存在
    os.makedirs(save_dir, exist_ok=True)
    
    # 获取文件扩展名
    parsed_url = urlparse(url)
    path = parsed_url.path
    extension = os.path.splitext(path)[1].lstrip('.').lower()
    if not extension or len(extension) > 10:  # 防止恶意扩展名
        extension = "jpg"  # 默认使用jpg
    
    # 生成唯一文件名
    save_path = generate_unique_filename(filename_prefix, extension, save_dir)
    
    headers = {}
    # 检查文件是否已存在，如果存在则生成新的文件名
    if os.path.exists(save_path):
        save_path = generate_unique_filename(filename_prefix, extension, save_dir)
    
    try:
        # 下载图片
        response = requests.get(url, headers=headers, stream=True, timeout=30)
        response.raise_for_status()
        
        # 以二进制写入模式打开文件
        with open(save_path, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
        
        return save_path
    except requests.exceptions.RequestException as e:
        # 如果下载失败，清理部分下载的文件
        if os.path.exists(save_path):
            os.remove(save_path) // ignore_security_alert
        raise Exception(f"图片下载失败: {str(e)}")
    except Exception as e:
        # 其他异常处理
        if os.path.exists(save_path):
            os.remove(save_path)
        raise Exception(f"下载图片时发生错误: {str(e)}")


def create_excel_with_images(results: list, temp_dir: str) -> str:
    """创建包含图片的Excel文件
    
    Args:
        results: 批量图片生成的结果列表
        temp_dir: 临时目录路径，用于保存下载的图片
    
    Returns:
        Excel文件的路径
    """
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "Image Generation Results"
    
    # 设置表头（使用英文表头避免编码问题）
    headers = ["Index", "Prompt", "Image"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 设置列宽
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 30
    
    # 遍历结果，插入数据和图片
    for row, item in enumerate(results, 2):
        # 设置行高
        ws.row_dimensions[row].height = 150
        
        # 序号
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        
        # 提示词
        ws.cell(row=row, column=2, value=item.get("question", ""))
        ws.cell(row=row, column=2).alignment = Alignment(vertical="center", wrap_text=True)
        
        # 图片
        if item.get("success", False):
            result_data = item.get("result", {})
            data = result_data.get("data", [])
            if data:
                try:
                    # 下载第一张图片
                    image_url = data[0].get("url", "")
                    if image_url:
                            # 下载图片，使用新的download_image函数接口
                        # 生成图片保存路径，使用新的命名方式
                        filename_prefix = f"image_{row-1}"
                        image_path = download_image(image_url, temp_dir, filename_prefix)
                        
                        # 插入图片到Excel
                        img = ExcelImage(image_path)
                        # 调整图片大小，保持比例
                        max_width = 200
                        max_height = 150
                        img.width = min(img.width, max_width)
                        img.height = min(img.height, max_height)
                        
                        # 设置图片位置
                        ws.add_image(img, f"C{row}")
                        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")
                except Exception as e:
                    ws.cell(row=row, column=3, value=f"Image insertion failed: {str(e)}")
                    ws.cell(row=row, column=3).alignment = Alignment(vertical="center", wrap_text=True)
        else:
            ws.cell(row=row, column=3, value="Failed to generate image")
            ws.cell(row=row, column=3).alignment = Alignment(vertical="center", wrap_text=True)
    
    # 保存Excel文件
    excel_path = os.path.join(temp_dir, "image_generation_results.xlsx")
    # 使用openpyxl的默认保存方法
    wb.save(excel_path)
    
    return excel_path

def process_uploaded_image(file_content: bytes) -> dict:
    """Process uploaded image and return image info"""
    try:
        # Generate unique ID
        image_id = str(uuid.uuid4())
        
        # Open image with PIL to validate and get info
        image = Image.open(BytesIO(file_content))
        
        # Convert to RGB if necessary
        if image.mode != 'RGB':
            image = image.convert('RGB')
        
        # Save the image to temporary directory
        file_path = os.path.join(UPLOAD_DIR, f"{image_id}.jpg")
        image.save(file_path, "JPEG", quality=95)
        
        # Store image info
        image_info = {
            'id': image_id,
            'file_path': file_path,
            'width': image.width,
            'height': image.height,
            'format': 'JPEG',
            'upload_time': time.time(),
            'base64_data': base64.b64encode(file_content).decode('utf-8')
        }
        
        UPLOADED_IMAGES[image_id] = image_info
        
        return image_info
    except Exception as e:
        raise Exception(f"Failed to process image: {str(e)}")

def get_image_base64(image_id: str) -> str:
    """Return data URL base64 string for saved JPEG image."""
    if image_id not in UPLOADED_IMAGES:
        raise Exception("Image not found")
    info = UPLOADED_IMAGES[image_id]
    file_path = info['file_path']
    if not os.path.exists(file_path):
        raise Exception("Saved image file not found")
    with open(file_path, 'rb') as f:
        jpeg_bytes = f.read()
    b64 = base64.b64encode(jpeg_bytes).decode('utf-8')
    # Use JPEG since we save processed file as JPEG
    return f"data:image/jpeg;base64,{b64}"


def build_messages(user_input: str, system_prompt: str):
    """Minimal messages builder to avoid streamlit dependency."""
    return [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_input},
    ]


def _calculate_qps(run_count: int) -> int:
    """Calculate QPS based on run_count rules."""
    if run_count <= 20:
        return 10
    else:
        # floor(run_count / 2)
        return min(run_count // 2, 40)



class EPChatPayload(BaseModel):
    input: str
    api_key: str | None = None
    model: str | None = None
    system_prompt: str | None = None
    thinking_mode: str | None = "disabled"
    reasoning_effort: str | None = "medium"
    temperature: float | None = 1.0
    top_p: float | None = 0.7
    max_tokens: int | None = 4094
    run_count: int | None = 1
    qps: int | None = 1

class NetworkQAPayload(BaseModel):
    input: str
    bearer_token: str | None = None
    bot_id: str | None = None
    system_prompt: str | None = None
    thinking: str | None = None
    run_count: int | None = 1


class PictureGenPayload(BaseModel):
    input: str
    api_key: str | None = None
    model: str | None = None
    size: str | None = None
    custom_size: str | None = None
    sequential: str | None = "disabled"  # disabled | enabled
    max_images: int | None = 1
    watermark: bool | None = False
    response_format: str | None = "url"  # url | b64_json
    input_mode: str | None = "text_to_image"  # text_to_image | image_to_image
    uploaded_image_ids: list[str] | None = None  # 上传的参考图片ID列表
    optimize_mode: str | None = "standard"  # standard | fast
    run_count: int | None = 1  # 并发请求次数
    qps: int | None = 1  # 并发QPS控制
    auto_download: bool | None = False  # 是否自动下载生成的图片


class BatchPictureGenPayload(BaseModel):
    questions: list[dict]  # 批量问题列表，每个问题包含question和可选的image_url字段
    api_key: str | None = None
    model: str | None = None
    size: str | None = None
    custom_size: str | None = None
    sequential: str | None = "disabled"  # disabled | enabled
    max_images: int | None = 1
    watermark: bool | None = False
    response_format: str | None = "url"  # url | b64_json
    optimize_mode: str | None = "standard"  # standard | fast
    input_mode: str | None = "text_to_image"  # text_to_image | image_to_image
    qps: int | None = 1  # 并发QPS控制
    download_excel: bool | None = False  # 是否需要下载Excel文件


class ExcelGenPayload(BaseModel):
    """生成Excel文件的请求模型"""
    results: list  # 图片生成的结果列表
    questions: list[str] | None = None  # 批量问题列表（可选）


class ExportResultsPayload(BaseModel):
    """导出结果的请求模型"""
    results: list  # 结果列表
    format: str = "json"  # json 或 csv
    type: str = "ep"  # ep 或 network_qa


class VideoGenPayload(BaseModel):
    input: str
    api_key: str | None = None
    model: str | None = None
    duration: int | None = 5
    ratio: str | None = None
    resolution: str | None = "720p"
    input_mode: str | None = "文生视频"
    run_count: int | None = 1
    qps: int | None = 1
    image_url: str | None = None
    image_urls: list[str] | None = None
    draft: bool | None = False
    generate_audio: bool | None = True
    return_last_frame: bool | None = False


class MultiModelPayload(BaseModel):
    input: str
    models: list[str]
    api_key: str | None = None
    system_prompt: str | None = None
    thinking_mode: str | None = "disabled"
    reasoning_effort: str | None = "medium"
    temperature: float | None = 1.0
    top_p: float | None = 0.7
    max_tokens: int | None = 4094


@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "ep_models_map": MODEL_EP_OPTIONS,
            "default_ep_model": EP_CONFIG.get("model_ep"),
            "default_thinking": EP_CONFIG.get("thinking_mode"),
            "default_temperature": EP_CONFIG.get("temperature"),
            "default_top_p": EP_CONFIG.get("top_p"),
            "default_max_tokens": EP_CONFIG.get("max_tokens"),
            # 图片与视频模型下拉选项
            "pic_models_map": PICTURE_GENERATION_CONFIG.get("models", {}),
            "default_pic_model": (next(iter(PICTURE_GENERATION_CONFIG.get("models", {}).values()), None)),
            "pic_sizes_map": PICTURE_GENERATION_CONFIG.get("sizes", {}),
            "default_pic_size": PICTURE_GENERATION_CONFIG.get("default_size", "4K"),
            "vid_models_map": VIDEO_GENERATION_CONFIG.get("models", {}),
            "default_vid_model": (next(iter(VIDEO_GENERATION_CONFIG.get("models", {}).values()), None)),
            # 即梦AI配置
            "jimeng_sizes_map": JIMENG_AI_CONFIG.get("sizes", {}),
            "default_jimeng_size": JIMENG_AI_CONFIG.get("default_size", 4194304),
            # 版本信息
            "full_version": FULL_VERSION,
            "version": VERSION,
            "app_name": APP_NAME,
        },
    )


def build_messages_with_image(user_input: str, system_prompt: str, image_base64: Optional[str] = None, mime_type: Optional[str] = None):
    """Build messages with optional image."""
    if image_base64 and mime_type:
        return [
            {"role": "system", "content": system_prompt},
            {
                "role": "user",
                "content": [
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:{mime_type};base64,{image_base64}"
                        }
                    },
                    {
                        "type": "text",
                        "text": user_input
                    }
                ]
            }
        ]
    else:
        return build_messages(user_input, system_prompt)


@app.post("/api/upload_image")
async def upload_image(file: UploadFile = File(...)):
    """Upload reference image for picture generation"""
    try:
        # Validate file type
        if not file.content_type.startswith('image/'):
            return JSONResponse({"error": "只能上传图片文件"}, status_code=400)
        
        # Read file content
        file_content = await file.read()
        
        # Process the image
        image_info = process_uploaded_image(file_content)
        
        # Clean up expired images
        cleanup_uploaded_images()
        
        return {
            "success": True,
            "image_id": image_info['id'],
            "width": image_info['width'],
            "height": image_info['height'],
            "message": "图片上传成功"
        }
    except Exception as e:
        return JSONResponse({"error": f"图片上传失败: {str(e)}"}, status_code=500)


@app.post("/api/delete_image")
async def delete_image(image_id: str = Form(...)):
    """Delete uploaded image by image_id"""
    try:
        if image_id not in UPLOADED_IMAGES:
            return JSONResponse({"error": "图片不存在"}, status_code=404)
        
        image_info = UPLOADED_IMAGES[image_id]
        file_path = image_info['file_path']
        
        # Delete the file from disk
        if os.path.exists(file_path):
            os.remove(file_path)
        
        # Remove from memory
        del UPLOADED_IMAGES[image_id]
        
        return {
            "success": True,
            "message": "图片删除成功"
        }
    except Exception as e:
        return JSONResponse({"error": f"图片删除失败: {str(e)}"}, status_code=500)


@app.post("/api/ep_chat")
def ep_chat(payload: EPChatPayload):
    # Existing endpoint - same as before, for backward compatibility
    try:
        api_key = payload.api_key or os.getenv("ARK_API_KEY") or EP_CONFIG.get("api_key")
        if not api_key:
            return JSONResponse({"error": "缺少 API Key，请在请求体或环境变量 ARK_API_KEY 中提供。"}, status_code=400)

        model = payload.model or EP_CONFIG.get("model_ep")
        system_prompt = payload.system_prompt or SYSTEM_PROMPT

        def single_call(inp: str):
            body = {
                "model": model,
                "messages": build_messages(inp, system_prompt),
                "temperature": payload.temperature or EP_CONFIG.get("temperature", 1.0),
                "top_p": payload.top_p or EP_CONFIG.get("top_p", 0.7),
                "max_tokens": payload.max_tokens or EP_CONFIG.get("max_tokens", 4094),
                "stream": False,
            }
            # 始终显式传递 thinking，避免服务端默认开启
            thinking_mode = payload.thinking_mode or 'disabled'
            body["thinking"] = {"type": thinking_mode}
            
            # 添加 reasoning 参数，仅当 thinking 为 enabled 时
            if thinking_mode == "enabled":
                body["reasoning"] = {"effort": payload.reasoning_effort or "medium"}

            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            }
            url = "https://ark.cn-beijing.volces.com/api/v3/chat/completions"
            r = requests.post(url, headers=headers, json=body, timeout=60)
            if r.status_code != 200:
                raise RuntimeError(f"EP API 状态码: {r.status_code}, 详情: {r.text}")
            resp = r.json()
            msg = resp.get("choices", [{}])[0].get("message", {})
            usage = resp.get("usage")
            return {
                "id": resp.get("id"),
                "content": msg.get("content", ""),
                "reasoning_content": msg.get("reasoning_content", None),
                "usage": usage,
                "system_prompt": system_prompt,
                "model": model,
                "thinking_mode": thinking_mode,
                "temperature": body["temperature"],
                "top_p": body["top_p"],
                "max_tokens": body["max_tokens"],
            }

        run_count = max(1, int(payload.run_count or 1))
        qps = _calculate_qps(run_count)

        if run_count == 1:
            result = single_call(payload.input)
            return {"results": [result]}
        else:
            from concurrent.futures import ThreadPoolExecutor, as_completed
            results = []
            with ThreadPoolExecutor(max_workers=qps) as ex:
                futures = [ex.submit(single_call, payload.input) for _ in range(run_count)]
                for future in as_completed(futures):
                    try:
                        results.append(future.result())
                    except Exception as e:
                        results.append({"id": None, "content": "", "reasoning_content": None, "error": str(e)})
            return {"results": results}
    except Exception as e:
        return JSONResponse({"error": f"EP 推理调用失败: {type(e).__name__}: {str(e)}"}, status_code=500)


@app.post("/api/ep_chat_with_image")
def ep_chat_with_image(
    input: str = Form(...),
    image: UploadFile = File(...),
    api_key: Optional[str] = Form(None),
    model: Optional[str] = Form(None),
    system_prompt: Optional[str] = Form(None),
    thinking_mode: Optional[str] = Form("disabled"),
    reasoning_effort: Optional[str] = Form("medium"),
    temperature: Optional[float] = Form(1.0),
    top_p: Optional[float] = Form(0.7),
    max_tokens: Optional[int] = Form(4094),
    run_count: Optional[int] = Form(1),
    qps: Optional[int] = Form(1),
):
    """EP推理接口，支持图片上传和并发请求"""
    try:
        # Get API key
        final_api_key = api_key or os.getenv("ARK_API_KEY") or EP_CONFIG.get("api_key")
        if not final_api_key:
            return JSONResponse({"error": "缺少 API Key"}, status_code=400)

        # Get model
        final_model = model or "doubao-1-5-vision-pro-32k-250115"
        final_system_prompt = system_prompt or SYSTEM_PROMPT
        
        # Ensure valid integers
        run_count = max(1, int(run_count or 1))
        qps = _calculate_qps(run_count)

        # Read and encode image
        image_bytes = image.file.read()
        image_base64 = base64.b64encode(image_bytes).decode('utf-8')
        mime_type = image.content_type or 'image/png'  # Default to png if content type not available

        # Build messages with image (reuse for all concurrent requests)
        messages = build_messages_with_image(input, final_system_prompt, image_base64, mime_type)

        def event_generator():
            out_q = queue.Queue()
            done = 0

            def worker(run_index: int):
                try:
                    start_time = time.time()
                    body = {
                        "model": final_model,
                        "messages": messages,
                        "temperature": temperature or EP_CONFIG.get("temperature", 1.0),
                        "top_p": top_p or EP_CONFIG.get("top_p", 0.7),
                        "max_tokens": max_tokens or EP_CONFIG.get("max_tokens", 4094),
                        "stream": True,
                        "stream_options": {"include_usage": True},
                    }
                    # 始终显式传递 thinking，避免服务端默认开启
                    body["thinking"] = {"type": thinking_mode or 'disabled'}
                    
                    # 添加 reasoning 参数，仅当 thinking 为 enabled 时
                    if (thinking_mode or 'disabled') == "enabled":
                        body["reasoning"] = {"effort": reasoning_effort or "medium"}

                    headers = {
                        "Authorization": f"Bearer {final_api_key}",
                        "Content-Type": "application/json",
                    }
                    url = "https://ark.cn-beijing.volces.com/api/v3/chat/completions"
                    # Increase timeout for image processing
                    with requests.post(url, headers=headers, json=body, stream=True, timeout=300) as r:
                        if r.status_code != 200:
                            try:
                                error_resp = r.json()
                                error_msg = error_resp.get("error", {}).get("message", r.text)
                            except:
                                error_msg = r.text
                            raise RuntimeError(f"API Error {r.status_code}: {error_msg}")

                        content_parts = []
                        reasoning_parts = []
                        usage = None
                        response_id = None
                        first_delta_time: float | None = None

                        for raw_line in r.iter_lines():
                            if not raw_line:
                                continue
                            line = raw_line.decode("utf-8").strip()
                            if line == "data:[DONE]":
                                break
                            
                            payload = None
                            if line.startswith("data:"):
                                try:
                                    payload = json.loads(line[5:])
                                except Exception:
                                    payload = None
                            else:
                                try:
                                    payload = json.loads(line)
                                except Exception:
                                    payload = None

                            if not payload:
                                continue

                            if not response_id and "id" in payload:
                                response_id = payload["id"]

                            # 标准格式：choices -> delta -> content/reasoning_content
                            sse_data = {}
                            if "choices" in payload and payload["choices"]:
                                for choice in payload["choices"]:
                                    delta = choice.get("delta", {})
                                    if "content" in delta and delta["content"]:
                                        content_parts.append(delta["content"])
                                        if first_delta_time is None:
                                            first_delta_time = time.time()
                                    if "reasoning_content" in delta and delta["reasoning_content"]:
                                        reasoning_parts.append(delta["reasoning_content"])
                                        if first_delta_time is None:
                                            first_delta_time = time.time()
                            
                            # 统计 tokens
                            if "usage" in payload and not usage:
                                u = payload["usage"]
                                if u is not None:
                                    usage = {
                                        "completion_tokens": u.get("completion_tokens"),
                                        "prompt_tokens": u.get("prompt_tokens"),
                                        "total_tokens": u.get("total_tokens"),
                                    }

                            # 发送增量更新
                            data_obj = {
                                "event": "content_delta",
                                "run_index": run_index,
                                "content": ''.join(content_parts),
                                "reasoning_content": ''.join(reasoning_parts) if reasoning_parts else None,
                            }
                            out_q.put(data_obj)

                    # 发送最终完成事件
                    final_data = {
                        "event": "final",
                        "run_index": run_index,
                        "id": response_id,
                        "content": ''.join(content_parts),
                        "reasoning_content": ''.join(reasoning_parts) if reasoning_parts else None,
                        "usage": usage,
                        "model": final_model,
                        "system_prompt": final_system_prompt,
                        "thinking_mode": thinking_mode,
                        "temperature": temperature or EP_CONFIG.get("temperature", 1.0),
                        "top_p": top_p or EP_CONFIG.get("top_p", 0.7),
                        "max_tokens": max_tokens or EP_CONFIG.get("max_tokens", 4094),
                        "first_token_time": (first_delta_time - start_time) if first_delta_time else None,
                        "total_time": time.time() - start_time,
                    }
                    out_q.put(final_data)

                except Exception as e:
                    error_data = {"event": "error", "run_index": run_index, "message": str(e)}
                    out_q.put(error_data)

            with ThreadPoolExecutor(max_workers=qps) as ex:
                futures = [ex.submit(worker, i) for i in range(run_count)]
                
                while done < run_count:
                    try:
                        item = out_q.get(timeout=0.5)
                        yield f"data: {json.dumps(item, ensure_ascii=False)}\n\n"
                        if item.get("event") in ("final", "error"):
                            done += 1
                    except queue.Empty:
                        yield "data: {\"event\": \"heartbeat\"}\n\n"

        return StreamingResponse(event_generator(), media_type="text/event-stream")

    except Exception as e:
        return JSONResponse({"error": f"EP 推理调用失败: {type(e).__name__}: {str(e)}"}, status_code=500)


@app.get("/api/ep_chat_stream")
def ep_chat_stream(
    input: str,
    api_key: Optional[str] = None,
    model: Optional[str] = None,
    system_prompt: Optional[str] = None,
    thinking_mode: Optional[str] = "disabled",
    reasoning_effort: Optional[str] = "medium",
    temperature: Optional[float] = None,
    top_p: Optional[float] = None,
    max_tokens: Optional[int] = None,
):
    """EP推理流式输出（SSE）。GET 参数用于简化前端 EventSource 使用。"""
    try:
        final_api_key = api_key or os.getenv("ARK_API_KEY") or EP_CONFIG.get("api_key")
        if not final_api_key:
            return JSONResponse({"error": "缺少 API Key"}, status_code=400)

        final_model = model or EP_CONFIG.get("model_ep")
        final_system_prompt = system_prompt or SYSTEM_PROMPT

        def event_generator():
            try:
                start_time = time.time()
                body = {
                    "model": final_model,
                    "messages": build_messages(input, final_system_prompt),
                    "temperature": temperature or EP_CONFIG.get("temperature"),
                    "top_p": top_p or EP_CONFIG.get("top_p"),
                    "max_tokens": max_tokens or EP_CONFIG.get("max_tokens"),
                    "stream": True,
                    "stream_options": {"include_usage": True},
                }
                # 始终显式传递 thinking，避免服务端默认开启
                body["thinking"] = {"type": (thinking_mode or 'disabled')}
                
                # 添加 reasoning 参数，仅当 thinking 为 enabled 时
                if (thinking_mode or 'disabled') == "enabled":
                    body["reasoning"] = {"effort": reasoning_effort or "medium"}

                headers = {
                    "Authorization": f"Bearer {final_api_key}",
                    "Content-Type": "application/json",
                }
                url = "https://ark.cn-beijing.volces.com/api/v3/chat/completions"
                with requests.post(url, headers=headers, json=body, stream=True, timeout=120) as r:
                    if r.status_code != 200:
                        try:
                            error_resp = r.json()
                            error_msg = error_resp.get("error", {}).get("message", r.text)
                        except:
                            error_msg = r.text
                        raise RuntimeError(f"API Error {r.status_code}: {error_msg}")

                    content_parts = []
                    reasoning_parts = []
                    usage = None
                    response_id = None
                    first_delta_time: float | None = None

                    for raw_line in r.iter_lines():
                        if not raw_line:
                            continue
                        line = raw_line.decode("utf-8").strip()
                        if line == "data:[DONE]":
                            break
                        # 兼容 'data:' 前缀和纯JSON行
                        payload = None
                        if line.startswith("data:"):
                            try:
                                payload = json.loads(line[5:])
                            except Exception:
                                payload = None
                        else:
                            try:
                                payload = json.loads(line)
                            except Exception:
                                payload = None

                        if not payload:
                            continue

                        if not response_id and "id" in payload:
                            response_id = payload["id"]

                        # 标准格式：choices -> delta -> content/reasoning_content
                        if "choices" in payload and payload["choices"]:
                            for choice in payload["choices"]:
                                delta = choice.get("delta", {})
                                if "content" in delta and delta["content"]:
                                    content_parts.append(delta["content"])
                                    if first_delta_time is None:
                                        first_delta_time = time.time()
                                if "reasoning_content" in delta and delta["reasoning_content"]:
                                    reasoning_parts.append(delta["reasoning_content"])
                                    if first_delta_time is None:
                                        first_delta_time = time.time()
                        # 统计 tokens
                        if "usage" in payload and not usage:
                            u = payload["usage"]
                            if u is not None:
                                usage = {
                                    "completion_tokens": u.get("completion_tokens"),
                                    "prompt_tokens": u.get("prompt_tokens"),
                                    "total_tokens": u.get("total_tokens"),
                                }

                        # 每次增量发送聚合后的内容
                        data_obj = {
                            "event": "content_delta",
                            "content": "".join(content_parts),
                            "reasoning_content": "".join(reasoning_parts) if reasoning_parts else None,
                        }
                        yield f"data: {json.dumps(data_obj, ensure_ascii=False)}\n\n"

                # 结束时发送最终信息
                final_obj = {
                    "event": "final",
                    "id": response_id,
                    "content": "".join(content_parts),
                    "reasoning_content": "".join(reasoning_parts) if reasoning_parts else None,
                    "usage": usage,
                    "model": final_model,
                    "system_prompt": final_system_prompt,
                    "thinking_mode": thinking_mode or 'disabled',
                    "temperature": temperature or EP_CONFIG.get("temperature"),
                    "top_p": top_p or EP_CONFIG.get("top_p"),
                    "max_tokens": max_tokens or EP_CONFIG.get("max_tokens"),
                    "first_token_time": (first_delta_time - start_time) if first_delta_time else None,
                    "total_time": time.time() - start_time,
                }
                yield f"data: {json.dumps(final_obj, ensure_ascii=False)}\n\n"
            except Exception as e:
                err = {"event": "error", "message": str(e)}
                yield f"data: {json.dumps(item, ensure_ascii=False)}\n\n"

        return StreamingResponse(event_generator(), media_type="text/event-stream")

    except Exception as e:
        return JSONResponse({"error": f"EP 推理调用失败: {type(e).__name__}: {str(e)}"}, status_code=500)


@app.post("/api/multi_model_chat")
def multi_model_chat(payload: MultiModelPayload):
    try:
        print(f"[DEBUG] 收到多模型对比请求: input={payload.input[:50]}..., models={payload.models}")
        
        api_key = payload.api_key or os.getenv("ARK_API_KEY") or EP_CONFIG.get("api_key")
        if not api_key:
            print("[DEBUG] API Key 缺失")
            return JSONResponse({"error": "缺少 API Key，请在请求体或环境变量 ARK_API_KEY 中提供。"}, status_code=400)

        system_prompt = payload.system_prompt or SYSTEM_PROMPT
        print(f"[DEBUG] 使用系统提示: {system_prompt[:50]}...")
        
        def worker(model_id: str, out_q: queue.Queue):
            try:
                print(f"[DEBUG] 开始处理模型 {model_id}")
                start_time = time.time()
                body = {
                    "model": model_id,
                    "messages": build_messages(payload.input, system_prompt),
                    "temperature": payload.temperature or EP_CONFIG.get("temperature", 1.0),
                    "top_p": payload.top_p or EP_CONFIG.get("top_p", 0.7),
                    "max_tokens": payload.max_tokens or EP_CONFIG.get("max_tokens", 4094),
                    "stream": True,
                    "stream_options": {"include_usage": True},
                }
                # 始终显式传递 thinking，避免服务端默认开启
                thinking_mode = payload.thinking_mode or 'disabled'
                body["thinking"] = {"type": thinking_mode}
                
                # 添加 reasoning 参数，仅当 thinking 为 enabled 时
                if thinking_mode == "enabled":
                    body["reasoning"] = {"effort": payload.reasoning_effort or "medium"}

                headers = {
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json",
                }
                url = "https://ark.cn-beijing.volces.com/api/v3/chat/completions"
                print(f"[DEBUG] 发送请求到模型 {model_id}")
                with requests.post(url, headers=headers, json=body, stream=True, timeout=300) as r:
                    print(f"[DEBUG] 模型 {model_id} 响应状态: {r.status_code}")
                    if r.status_code != 200:
                        try:
                            error_resp = r.json()
                            error_msg = error_resp.get("error", {}).get("message", r.text)
                        except:
                            error_msg = r.text
                        raise RuntimeError(f"API Error {r.status_code}: {error_msg}")

                    content_parts = []
                    reasoning_parts = []
                    usage = None
                    response_id = None
                    first_delta_time: float | None = None

                    for raw_line in r.iter_lines():
                        if not raw_line:
                            continue
                        line = raw_line.decode("utf-8").strip()
                        if line == "data:[DONE]":
                            break
                        
                        payload_data = None
                        if line.startswith("data:"):
                            try:
                                payload_data = json.loads(line[5:])
                            except Exception:
                                payload_data = None
                        else:
                            try:
                                payload_data = json.loads(line)
                            except Exception:
                                payload_data = None

                        if not payload_data:
                            continue

                        if not response_id and "id" in payload_data:
                            response_id = payload_data["id"]

                        # 标准格式：choices -> delta -> content/reasoning_content
                        sse_data = {}
                        if "choices" in payload_data and payload_data["choices"]:
                            for choice in payload_data["choices"]:
                                delta = choice.get("delta", {})
                                if "content" in delta and delta["content"]:
                                    content_parts.append(delta["content"])
                                    if first_delta_time is None:
                                        first_delta_time = time.time()
                                if "reasoning_content" in delta and delta["reasoning_content"]:
                                    reasoning_parts.append(delta["reasoning_content"])
                                    if first_delta_time is None:
                                        first_delta_time = time.time()
                        
                        # 统计 tokens
                        if "usage" in payload_data and not usage:
                            u = payload_data["usage"]
                            if u is not None:
                                usage = {
                                    "completion_tokens": u.get("completion_tokens"),
                                    "prompt_tokens": u.get("prompt_tokens"),
                                    "total_tokens": u.get("total_tokens"),
                                }

                        # 发送增量更新
                        data_obj = {
                            "event": "content_delta",
                            "model_id": model_id,
                            "content": ''.join(content_parts),
                            "reasoning_content": ''.join(reasoning_parts) if reasoning_parts else None,
                        }
                        print(f"[DEBUG] 模型 {model_id} 发送增量更新: {len(content_parts)} 个字符")
                        out_q.put(data_obj)

                # 发送最终完成事件
                final_data = {
                    "event": "final",
                    "model_id": model_id,
                    "id": response_id,
                    "content": ''.join(content_parts),
                    "reasoning_content": ''.join(reasoning_parts) if reasoning_parts else None,
                    "usage": usage,
                    "model": model_id,
                    "system_prompt": system_prompt,
                    "thinking_mode": thinking_mode,
                    "temperature": body["temperature"],
                    "top_p": body["top_p"],
                    "max_tokens": body["max_tokens"],
                    "first_token_time": (first_delta_time - start_time) if first_delta_time else None,
                    "total_time": time.time() - start_time,
                }
                print(f"[DEBUG] 模型 {model_id} 完成: {len(content_parts)} 个字符")
                out_q.put(final_data)

            except Exception as e:
                print(f"[DEBUG] 模型 {model_id} 错误: {str(e)}")
                error_data = {"event": "error", "model_id": model_id, "message": str(e)}
                out_q.put(error_data)

        def event_generator():
            out_q = queue.Queue()
            done = 0
            total_models = len(payload.models)
            print(f"[DEBUG] 开始处理 {total_models} 个模型")

            with ThreadPoolExecutor(max_workers=len(payload.models)) as ex:
                futures = [ex.submit(worker, model, out_q) for model in payload.models]
                
                while done < total_models:
                    try:
                        item = out_q.get(timeout=0.5)
                        print(f"[DEBUG] 发送事件: {item['event']} for model {item.get('model_id')}")
                        yield f"data: {json.dumps(item, ensure_ascii=False)}\n\n"
                        if item.get("event") in ("final", "error"):
                            done += 1
                            print(f"[DEBUG] 完成模型数: {done}/{total_models}")
                    except queue.Empty:
                        yield "data: {\"event\": \"heartbeat\"}\n\n"

            print(f"[DEBUG] 所有模型处理完成")

    except Exception as e:
        print(f"[DEBUG] 多模型对比调用失败: {str(e)}")
        return JSONResponse({"error": f"多模型对比调用失败: {type(e).__name__}: {str(e)}"}, status_code=500)

    # 返回流式响应
    return StreamingResponse(event_generator(), media_type="text/plain")


@app.get("/api/ep_chat_stream_multi")
def ep_chat_stream_multi(
    input: str,
    run_count: int = 1,
    qps: int = 1,
    api_key: Optional[str] = None,
    model: Optional[str] = None,
    system_prompt: Optional[str] = None,
    thinking_mode: Optional[str] = "disabled",
    reasoning_effort: Optional[str] = "medium",
    temperature: Optional[float] = None,
    top_p: Optional[float] = None,
    max_tokens: Optional[int] = None,
):
    """EP推理多路并发流式输出（SSE聚合）。单连接中复用多个并发流，事件携带 run_index。"""
    try:
        final_api_key = api_key or os.getenv("ARK_API_KEY") or EP_CONFIG.get("api_key")
        if not final_api_key:
            return JSONResponse({"error": "缺少 API Key"}, status_code=400)

        final_model = model or EP_CONFIG.get("model_ep")
        final_system_prompt = system_prompt or SYSTEM_PROMPT
        run_count = max(1, int(run_count))
        qps = _calculate_qps(run_count)

        def event_generator():
            out_q = queue.Queue()
            done = 0

            def worker(run_index: int):
                try:
                    start_time = time.time()
                    body = {
                        "model": final_model,
                        "messages": build_messages(input, final_system_prompt),
                        "temperature": temperature or EP_CONFIG.get("temperature", 1.0),
                        "top_p": top_p or EP_CONFIG.get("top_p", 0.7),
                        "max_tokens": max_tokens or EP_CONFIG.get("max_tokens", 4094),
                        "stream": True,
                        "stream_options": {"include_usage": True},
                    }
                    # 始终显式传递 thinking，避免服务端默认开启
                    body["thinking"] = {"type": (thinking_mode or 'disabled')}
                    
                    # 添加 reasoning 参数，仅当 thinking 为 enabled 时
                    if (thinking_mode or 'disabled') == "enabled":
                        body["reasoning"] = {"effort": reasoning_effort or "medium"}

                    headers = {
                        "Authorization": f"Bearer {final_api_key}",
                        "Content-Type": "application/json",
                    }
                    url = "https://ark.cn-beijing.volces.com/api/v3/chat/completions"
                    with requests.post(url, headers=headers, json=body, stream=True, timeout=120) as r:
                        if r.status_code != 200:
                            try:
                                error_resp = r.json()
                                error_msg = error_resp.get("error", {}).get("message", r.text)
                            except:
                                error_msg = r.text
                            raise RuntimeError(f"API Error {r.status_code}: {error_msg}")

                        content_parts = []
                        reasoning_parts = []
                        usage = None
                        response_id = None
                        first_delta_time: float | None = None

                        for raw_line in r.iter_lines():
                            if not raw_line:
                                continue
                            line = raw_line.decode("utf-8").strip()
                            if line == "data:[DONE]":
                                break
                            
                            payload = None
                            if line.startswith("data:"):
                                try:
                                    payload = json.loads(line[5:])
                                except Exception:
                                    payload = None
                            else:
                                try:
                                    payload = json.loads(line)
                                except Exception:
                                    payload = None

                            if not payload:
                                continue

                            if not response_id and "id" in payload:
                                response_id = payload["id"]

                            if "choices" in payload and payload["choices"]:
                                for choice in payload["choices"]:
                                    delta = choice.get("delta", {})
                                    if "content" in delta and delta["content"]:
                                        content_parts.append(delta["content"])
                                        if first_delta_time is None:
                                            first_delta_time = time.time()
                                    if "reasoning_content" in delta and delta["reasoning_content"]:
                                        reasoning_parts.append(delta["reasoning_content"])
                                        if first_delta_time is None:
                                            first_delta_time = time.time()

                            if "usage" in payload and not usage:
                                u = payload["usage"]
                                if u is not None:
                                    usage = {
                                        "completion_tokens": u.get("completion_tokens"),
                                        "prompt_tokens": u.get("prompt_tokens"),
                                        "total_tokens": u.get("total_tokens"),
                                    }

                            data_obj = {
                                "event": "content_delta",
                                "run_index": run_index,
                                "content": "".join(content_parts),
                                "reasoning_content": "".join(reasoning_parts) if reasoning_parts else None,
                            }
                            out_q.put(data_obj)

                    final_obj = {
                        "event": "final",
                        "run_index": run_index,
                        "id": response_id,
                        "content": "".join(content_parts),
                        "reasoning_content": "".join(reasoning_parts) if reasoning_parts else None,
                        "usage": usage,
                        "model": final_model,
                        "system_prompt": final_system_prompt,
                        "thinking_mode": thinking_mode or 'disabled',
                        "temperature": temperature or EP_CONFIG.get("temperature"),
                        "top_p": top_p or EP_CONFIG.get("top_p"),
                        "max_tokens": max_tokens or EP_CONFIG.get("max_tokens"),
                        "first_token_time": (first_delta_time - start_time) if first_delta_time else None,
                        "total_time": time.time() - start_time,
                    }
                    out_q.put(final_obj)
                except Exception as e:
                    out_q.put({"event": "error", "run_index": run_index, "message": str(e)})

            with ThreadPoolExecutor(max_workers=qps) as ex:
                futures = [ex.submit(worker, i) for i in range(run_count)]

                while done < run_count:
                    try:
                        item = out_q.get(timeout=0.5)
                        yield f"data: {json.dumps(item, ensure_ascii=False)}\n\n"
                        if item.get("event") in ("final", "error"):
                            done += 1
                    except queue.Empty:
                        # 发送心跳以保持连接活跃（可选）
                        yield "data: {\"event\": \"heartbeat\"}\n\n"

        return StreamingResponse(event_generator(), media_type="text/event-stream")
    except Exception as e:
        return JSONResponse({"error": f"EP 流式并发调用失败: {type(e).__name__}: {str(e)}"}, status_code=500)


@app.post("/api/network_qa_chat")
def network_qa_chat(payload: NetworkQAPayload):
    try:
        start_time = time.time()
        token = payload.bearer_token or NETWORK_QA_CONFIG.get("bearer_token")
        bot_id = payload.bot_id or NETWORK_QA_CONFIG.get("bot_id")
        system_prompt = payload.system_prompt or NETWORK_QA_CONFIG.get("system_prompt")

        if not token:
            return JSONResponse({"error": "缺少 Bearer Token"}, status_code=400)
        if not bot_id:
            return JSONResponse({"error": "缺少 bot_id"}, status_code=400)

        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json; charset=utf-8",
        }
        data = {
            "bot_id": bot_id,
            "messages": build_messages(payload.input, system_prompt),
            "stream": True,
        }
        # 根据传入的thinking参数值设置模型
        if payload.thinking == "thinking" or payload.thinking is True:
            data["model"] = "thinking"
        elif payload.thinking == "auto_thinking":
            data["model"] = "auto_thinking"

        url = NETWORK_QA_CONFIG.get("base_url")
        r = requests.post(url, headers=headers, json=data, stream=True, timeout=60)
        if r.status_code != 200:
            return JSONResponse({"error": f"API 状态码: {r.status_code}"}, status_code=r.status_code)

        content_parts: list[str] = []
        reasoning_parts: list[str] = []
        usage = None
        response_id = None
        first_delta_time: float | None = None

        for raw_line in r.iter_lines():
            if not raw_line:
                continue
            line = raw_line.decode("utf-8").strip()
            if line == "data:[DONE]":
                break
            if not line.startswith("data:"):
                continue
            try:
                chunk_data = json.loads(line[5:])
                if not response_id and "id" in chunk_data:
                    response_id = chunk_data["id"]

                # 标准格式：choices -> delta -> content
                if "choices" in chunk_data and chunk_data["choices"]:
                    for choice in chunk_data["choices"]:
                        delta = choice.get("delta", {})
                        if "content" in delta and delta["content"]:
                            content_parts.append(delta["content"])
                            if first_delta_time is None:
                                first_delta_time = time.time()
                        if "reasoning_content" in delta and delta["reasoning_content"]:
                            reasoning_parts.append(delta["reasoning_content"])
                # 备用格式：content/text/message.content
                elif "content" in chunk_data and chunk_data["content"]:
                    content_parts.append(chunk_data["content"])
                    if first_delta_time is None:
                        first_delta_time = time.time()
                elif "text" in chunk_data and chunk_data["text"]:
                    content_parts.append(chunk_data["text"])
                    if first_delta_time is None:
                        first_delta_time = time.time()
                elif "message" in chunk_data and chunk_data["message"].get("content"):
                    content_parts.append(chunk_data["message"]["content"])
                    if first_delta_time is None:
                        first_delta_time = time.time()

                # 统计 tokens
                if "usage" in chunk_data:
                    u = chunk_data["usage"]
                    usage = {
                        "completion_tokens": u.get("completion_tokens"),
                        "prompt_tokens": u.get("prompt_tokens"),
                        "total_tokens": u.get("total_tokens"),
                    }
                elif "bot_usage" in chunk_data and "model_usage" in chunk_data["bot_usage"]:
                    models = chunk_data["bot_usage"]["model_usage"]
                    usage = {
                        "completion_tokens": sum(m.get("completion_tokens", 0) for m in models),
                        "prompt_tokens": sum(m.get("prompt_tokens", 0) for m in models),
                        "total_tokens": sum(m.get("total_tokens", 0) for m in models),
                    }
            except Exception:
                # 忽略解析错误，继续累积
                continue

        final_content = "".join(content_parts)
        return {
            "id": response_id,
            "content": final_content,
            "reasoning_content": "".join(reasoning_parts),
            "usage": usage,
            "first_token_time": (first_delta_time - start_time) if first_delta_time else None,
            "total_time": time.time() - start_time,
        }
    except Exception as e:
        return JSONResponse({"error": f"联网问答调用失败: {type(e).__name__}: {str(e)}"}, status_code=500)


@app.get("/api/network_qa_chat_stream")
def network_qa_chat_stream(
    input: str,
    bearer_token: Optional[str] = None,
    bot_id: Optional[str] = None,
    system_prompt: Optional[str] = None,
    thinking: Optional[str] = "false",
):
    """联网问答流式输出（SSE）。"""
    try:
        token = bearer_token or NETWORK_QA_CONFIG.get("bearer_token")
        final_bot_id = bot_id or NETWORK_QA_CONFIG.get("bot_id")
        final_system_prompt = system_prompt or NETWORK_QA_CONFIG.get("system_prompt")

        if not token:
            return JSONResponse({"error": "缺少 Bearer Token"}, status_code=400)
        if not final_bot_id:
            return JSONResponse({"error": "缺少 bot_id"}, status_code=400)

        def event_generator():
            try:
                start_time = time.time()
                headers = {
                    "Authorization": f"Bearer {token}",
                    "Content-Type": "application/json; charset=utf-8",
                }
                body = {
                    "bot_id": final_bot_id,
                    "messages": build_messages(input, final_system_prompt),
                    "stream": True,
                }
                # 根据传入的thinking参数值设置模型
                if thinking == "thinking" or thinking == "true":
                    body["model"] = "thinking"
                elif thinking == "auto_thinking":
                    body["model"] = "auto_thinking"

                url = NETWORK_QA_CONFIG.get("base_url")
                with requests.post(url, headers=headers, json=body, stream=True, timeout=120) as r:
                    if r.status_code != 200:
                        raise RuntimeError(f"API 状态码: {r.status_code}, 详情: {r.text}")

                    content_parts = []
                    reasoning_parts = []
                    usage = None
                    response_id = None
                    first_delta_time: float | None = None

                    for raw_line in r.iter_lines():
                        if not raw_line:
                            continue
                        line = raw_line.decode("utf-8").strip()
                        if line == "data:[DONE]":
                            break
                        if not line.startswith("data:"):
                            continue
                        
                        try:
                            chunk_data = json.loads(line[5:])
                            # 日志：打印完整的API响应块
                            print(f"Network QA API Response Chunk: {json.dumps(chunk_data, ensure_ascii=False)}")
                        except Exception:
                            continue

                        if not response_id and "id" in chunk_data:
                            response_id = chunk_data["id"]

                        # 提取内容
                        delta_content = ""
                        delta_reasoning = ""
                        if "choices" in chunk_data and chunk_data["choices"]:
                            for choice in chunk_data["choices"]:
                                delta = choice.get("delta", {})
                                if "content" in delta and delta["content"]:
                                    delta_content = delta["content"]
                                if "reasoning_content" in delta and delta["reasoning_content"]:
                                    delta_reasoning = delta["reasoning_content"]
                        elif "content" in chunk_data and chunk_data["content"]:
                            delta_content = chunk_data["content"]
                        elif "text" in chunk_data and chunk_data["text"]:
                            delta_content = chunk_data["text"]
                        elif "message" in chunk_data and chunk_data["message"].get("content"):
                            delta_content = chunk_data["message"]["content"]

                        if delta_reasoning:
                            reasoning_parts.append(delta_reasoning)

                        if delta_content:
                            if first_delta_time is None:
                                first_delta_time = time.time()

                        # 统计 tokens
                        if "usage" in chunk_data:
                            u = chunk_data["usage"]
                            usage = {
                                "completion_tokens": u.get("completion_tokens"),
                                "prompt_tokens": u.get("prompt_tokens"),
                                "total_tokens": u.get("total_tokens"),
                            }
                        elif "bot_usage" in chunk_data and "model_usage" in chunk_data["bot_usage"]:
                            models = chunk_data["bot_usage"]["model_usage"]
                            usage = {
                                "completion_tokens": sum(m.get("completion_tokens", 0) for m in models),
                                "prompt_tokens": sum(m.get("prompt_tokens", 0) for m in models),
                                "total_tokens": sum(m.get("total_tokens", 0) for m in models),
                            }

                        # 更新内容部分
                        if delta_content:
                            content_parts.append(delta_content)
                        # 发送增量
                        data_obj = {
                            "event": "content_delta",
                            "content": "".join(content_parts),
                            "reasoning_content": "".join(reasoning_parts)
                        }
                        yield f"data: {json.dumps(data_obj, ensure_ascii=False)}\n\n"

                # 最终事件
                final_obj = {
                    "event": "final",
                    "id": response_id,
                    "content": "".join(content_parts),
                    "reasoning_content": "".join(reasoning_parts),
                    "usage": usage,
                    "first_token_time": (first_delta_time - start_time) if first_delta_time else None,
                    "total_time": time.time() - start_time,
                }
                yield f"data: {json.dumps(final_obj, ensure_ascii=False)}\n\n"

            except Exception as e:
                err = {"event": "error", "message": str(e)}
                yield f"data: {json.dumps(err, ensure_ascii=False)}\n\n"

        return StreamingResponse(event_generator(), media_type="text/event-stream")
    except Exception as e:
        return JSONResponse({"error": f"联网问答流式调用失败: {type(e).__name__}: {str(e)}"}, status_code=500)


@app.get("/api/network_qa_chat_stream_multi")
def network_qa_chat_stream_multi(
    input: str,
    run_count: int = 1,
    bearer_token: Optional[str] = None,
    bot_id: Optional[str] = None,
    system_prompt: Optional[str] = None,
    thinking: Optional[str] = "false",
):
    """联网问答多路并发流式输出（SSE聚合）。"""
    try:
        token = bearer_token or NETWORK_QA_CONFIG.get("bearer_token")
        final_bot_id = bot_id or NETWORK_QA_CONFIG.get("bot_id")
        final_system_prompt = system_prompt or NETWORK_QA_CONFIG.get("system_prompt")
        
        if not token:
            return JSONResponse({"error": "缺少 Bearer Token"}, status_code=400)
        if not final_bot_id:
            return JSONResponse({"error": "缺少 bot_id"}, status_code=400)

        run_count = max(1, int(run_count))
        qps = _calculate_qps(run_count)

        def event_generator():
            out_q = queue.Queue()
            done = 0

            def worker(run_index: int):
                try:
                    start_time = time.time()
                    headers = {
                        "Authorization": f"Bearer {token}",
                        "Content-Type": "application/json; charset=utf-8",
                    }
                    body = {
                        "bot_id": final_bot_id,
                        "messages": build_messages(input, final_system_prompt),
                        "stream": True,
                    }
                    # 根据传入的thinking参数值设置模型
                    if thinking == "thinking" or thinking == "true":
                        body["model"] = "thinking"
                    elif thinking == "auto_thinking":
                        body["model"] = "auto_thinking"

                    url = NETWORK_QA_CONFIG.get("base_url")
                    with requests.post(url, headers=headers, json=body, stream=True, timeout=120) as r:
                        if r.status_code != 200:
                            raise RuntimeError(f"API 状态码: {r.status_code}, 详情: {r.text}")

                        content_parts = []
                        reasoning_parts = []
                        usage = None
                        response_id = None
                        first_delta_time: float | None = None

                        for raw_line in r.iter_lines():
                            if not raw_line:
                                continue
                            line = raw_line.decode("utf-8").strip()
                            if line == "data:[DONE]":
                                break
                            if not line.startswith("data:"):
                                continue
                            
                            try:
                                chunk_data = json.loads(line[5:])
                            except Exception:
                                continue

                            if not response_id and "id" in chunk_data:
                                response_id = chunk_data["id"]

                            delta_content = ""
                            delta_reasoning = ""
                            if "choices" in chunk_data and chunk_data["choices"]:
                                for choice in chunk_data["choices"]:
                                    delta = choice.get("delta", {})
                                    if "content" in delta and delta["content"]:
                                        delta_content = delta["content"]
                                    if "reasoning_content" in delta and delta["reasoning_content"]:
                                        delta_reasoning = delta["reasoning_content"]
                            elif "content" in chunk_data and chunk_data["content"]:
                                delta_content = chunk_data["content"]
                                delta_reasoning = chunk_data.get("reasoning_content", "")
                            elif "text" in chunk_data and chunk_data["text"]:
                                delta_content = chunk_data["text"]
                                delta_reasoning = chunk_data.get("reasoning_content", "")
                            elif "message" in chunk_data and chunk_data["message"].get("content"):
                                delta_content = chunk_data["message"]["content"]
                                delta_reasoning = chunk_data["message"].get("reasoning_content", "")

                            # 累积内容
                            if delta_content:
                                content_parts.append(delta_content)
                                if first_delta_time is None:
                                    first_delta_time = time.time()
                            if delta_reasoning:
                                reasoning_parts.append(delta_reasoning)
                                if first_delta_time is None:
                                    first_delta_time = time.time()

                            # 发送增量更新到队列
                            if delta_content or delta_reasoning:
                                out_q.put({
                                    "event": "content_delta",
                                    "run_index": run_index,
                                    "content": "".join(content_parts),
                                    "reasoning_content": "".join(reasoning_parts) if reasoning_parts else None,
                                })

                            if "usage" in chunk_data:
                                u = chunk_data["usage"]
                                usage = {
                                    "completion_tokens": u.get("completion_tokens"),
                                    "prompt_tokens": u.get("prompt_tokens"),
                                    "total_tokens": u.get("total_tokens"),
                                }
                            elif "bot_usage" in chunk_data and "model_usage" in chunk_data["bot_usage"]:
                                models = chunk_data["bot_usage"]["model_usage"]
                                usage = {
                                    "completion_tokens": sum(m.get("completion_tokens", 0) for m in models),
                                    "prompt_tokens": sum(m.get("prompt_tokens", 0) for m in models),
                                    "total_tokens": sum(m.get("total_tokens", 0) for m in models),
                                }

                        # 发送最终完成事件
                        out_q.put({
                            "event": "final",
                            "run_index": run_index,
                            "id": response_id,
                            "content": "".join(content_parts),
                            "reasoning_content": "".join(reasoning_parts) if reasoning_parts else None,
                            "usage": usage,
                            "first_token_time": (first_delta_time - start_time) if first_delta_time else None,
                            "total_time": time.time() - start_time,
                        })

                except Exception as e:
                    out_q.put({"event": "error", "run_index": run_index, "message": str(e)})

            with ThreadPoolExecutor(max_workers=qps) as ex:
                futures = [ex.submit(worker, i) for i in range(run_count)]

                while done < run_count:
                    try:
                        item = out_q.get(timeout=0.5)
                        yield f"data: {json.dumps(item, ensure_ascii=False)}\n\n"
                        if item.get("event") in ("final", "error"):
                            done += 1
                    except queue.Empty:
                        yield "data: {\"event\": \"heartbeat\"}\n\n"

        return StreamingResponse(event_generator(), media_type="text/event-stream")
    except Exception as e:
        return JSONResponse({"error": f"联网问答流式并发调用失败: {type(e).__name__}: {str(e)}"}, status_code=500)


@app.post("/api/picture_generate")
def picture_generate(payload: PictureGenPayload):
    try:
        api_key = payload.api_key or PICTURE_GENERATION_CONFIG.get("api_key")
        if not api_key:
            return JSONResponse({"error": "缺少 API Key"}, status_code=400)

        model = payload.model or list(PICTURE_GENERATION_CONFIG.get("models", {}).values())[0]
        # 使用自定义尺寸或预设尺寸
        sizes_map = PICTURE_GENERATION_CONFIG.get("sizes", {})
        size_value = None
        if payload.size and payload.size in sizes_map:
            size_value = sizes_map[payload.size]
        elif payload.custom_size:
            size_value = payload.custom_size
        else:
            # 根据输入模式选择默认尺寸
            if payload.input_mode == "image_to_image":
                # 图生图默认4K
                size_value = sizes_map.get(PICTURE_GENERATION_CONFIG.get("default_i2i_size", "4K"), "4K")
            else:
                # 文生图默认16:9
                size_value = sizes_map.get(PICTURE_GENERATION_CONFIG.get("default_size", "16:9 (2560x1440)"), "2560x1440")

        # 处理并发参数
        run_count = max(1, int(payload.run_count or 1))
        # 使用config.py中的默认QPS值，不再从前端获取
        qps = PICTURE_GENERATION_CONFIG.get('picture_generation_qps', 2)

        def single_picture_call():
            """单次图片生成调用"""
            start_time = time.time()  # 记录单张图片开始时间
            
            # 构建请求体
            body = {
                "model": model,
                "prompt": payload.input,
                "size": size_value,
                "sequential_image_generation": payload.sequential or "disabled",
                "watermark": bool(payload.watermark),
                "response_format": payload.response_format or "url",
            }
            # 提示词优化选项：支持doubao-seedream-4.0系列和doubao-seedream-4.5系列
            if "doubao-seedream-4.0" in model or "doubao-seedream-4-0" in model or "doubao-seedream-4.5" in model:
                body["optimize_prompt_options"] = {
                    "mode": payload.optimize_mode or "standard"
                }
            
            # 处理图生图模式：设置 Ark API 所需的 image 参数
            if payload.input_mode == "image_to_image" and payload.uploaded_image_ids:
                for image_id in payload.uploaded_image_ids:
                    try:
                        data_url = get_image_base64(image_id)
                        body["image"] = data_url  # API 要求单个 image 参数，只使用第一张图
                        break
                    except Exception as e:
                        print(f"Warning: Failed to get image {image_id}: {e}")
            
            if body["sequential_image_generation"] == "auto":
                body["sequential_image_generation_options"] = {"max_images": payload.max_images or 1}

            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
                "User-Agent": "Seedream-Image-Generator/1.0",
            }

            url = PICTURE_GENERATION_CONFIG.get("api_endpoint")
            r = requests.post(url, headers=headers, json=body, timeout=120)  # 增加超时时间到120秒
            if r.status_code != 200:
                raise RuntimeError(f"图片生成接口状态码: {r.status_code}, 详情: {r.text}")
            

            result = r.json()
            # 确保size字段存在于每个图片对象中
            for image in result.get("data", []):
                if "size" not in image:
                    image["size"] = size_value
            # 从图片URL中提取request_id
            request_id = ""
            if result.get("data") and len(result["data"]) > 0:
                image_url = result["data"][0].get("url", "")
                if image_url:
                    # 从URL中提取request_id，格式类似：0217664856699576246fa05d035050efc9e62170c29c5f2181e1b_0.jpeg
                    import re
                    match = re.search(r'/(\w+)_0\.jpeg', image_url)
                    if match:
                        request_id = match.group(1)
            # 添加request_id到结果中
            result["request_id"] = request_id
            # 添加单张图片的生成耗时
            result["generation_time"] = round(time.time() - start_time, 2)
            # 添加输入模式信息
            result["input_mode"] = payload.input_mode or "text_to_image"
            return result

        # 单次请求直接返回
        if run_count == 1:
            result = single_picture_call()
            
            # 如果需要自动下载图片
            if payload.auto_download and result.get("data"):
                try:
                    # 下载第一张图片
                    image_url = result["data"][0].get("url", "")
                    if image_url:
                        # 创建用于保存自动下载图片的目录
                        auto_download_dir = os.path.join(os.getcwd(), "auto_downloaded_images")
                        os.makedirs(auto_download_dir, exist_ok=True)
                        
                        # 下载图片
                        download_image(image_url, auto_download_dir, "auto_gen_image")
                        # 将下载状态添加到结果中
                        result["auto_download"] = {"status": "success", "message": "图片已自动下载到本地目录"}
                except Exception as e:
                    # 如果下载失败，将错误信息添加到结果中，但不影响图片生成结果
                    result["auto_download"] = {"status": "failed", "message": f"图片自动下载失败: {str(e)}"}
            
            return {"results": [result]}
        
        # 并发请求
        else:
            from concurrent.futures import ThreadPoolExecutor, as_completed
            results = []
            
            with ThreadPoolExecutor(max_workers=qps) as ex:
                futures = [ex.submit(single_picture_call) for _ in range(run_count)]
                for future in as_completed(futures):
                    try:
                        result = future.result()
                        
                        # 如果需要自动下载图片
                        if payload.auto_download and result.get("data"):
                            try:
                                # 下载第一张图片
                                image_url = result["data"][0].get("url", "")
                                if image_url:
                                    # 创建用于保存自动下载图片的目录
                                    auto_download_dir = os.path.join(os.getcwd(), "auto_downloaded_images")
                                    os.makedirs(auto_download_dir, exist_ok=True)
                                    
                                    # 下载图片
                                    download_image(image_url, auto_download_dir, "auto_gen_image")
                                    # 将下载状态添加到结果中
                                    result["auto_download"] = {"status": "success", "message": "图片已自动下载到本地目录"}
                            except Exception as e:
                                # 如果下载失败，将错误信息添加到结果中，但不影响图片生成结果
                                result["auto_download"] = {"status": "failed", "message": f"图片自动下载失败: {str(e)}"}
                        
                        results.append(result)
                    except Exception as e:
                        results.append({"error": str(e)})
            
            return {"results": results}
            
    except Exception as e:
        return JSONResponse({"error": f"图片生成调用失败: {type(e).__name__}: {str(e)}"}, status_code=500)


@app.post("/api/picture_generate_batch")
def picture_generate_batch(payload: BatchPictureGenPayload):
    try:
        api_key = payload.api_key or PICTURE_GENERATION_CONFIG.get("api_key")
        if not api_key:
            return JSONResponse({"error": "缺少 API Key"}, status_code=400)

        model = payload.model or list(PICTURE_GENERATION_CONFIG.get("models", {}).values())[0]
        # 使用自定义尺寸或预设尺寸
        sizes_map = PICTURE_GENERATION_CONFIG.get("sizes", {})
        size_value = None
        if payload.size and payload.size in sizes_map:
            size_value = sizes_map[payload.size]
        elif payload.custom_size:
            size_value = payload.custom_size
        else:
            # 根据输入模式选择默认尺寸
            if payload.input_mode == "image_to_image":
                # 图生图默认4K
                size_value = sizes_map.get(PICTURE_GENERATION_CONFIG.get("default_i2i_size", "4K"), "4K")
            else:
                # 文生图默认16:9
                size_value = sizes_map.get(PICTURE_GENERATION_CONFIG.get("default_size", "16:9 (2560x1440)"), "2560x1440")

        # 处理并发参数
        qps = PICTURE_GENERATION_CONFIG.get('picture_generation_qps', 2)
        
        def single_picture_call(question_data):
            """单次图片生成调用，包含重试机制"""
            start_time = time.time()  # 记录单张图片开始时间
            max_retries = 3
            retry_delay = 2  # 初始重试延迟秒数
            
            # 提取问题文本和参考图片信息
            if isinstance(question_data, dict):
                question = question_data.get("question", "")
                image_url = question_data.get("image_url")
                image_base64 = question_data.get("image_base64")
                uploaded_image_id = question_data.get("uploaded_image_id")
            else:
                question = question_data
                image_url = None
                image_base64 = None
                uploaded_image_id = None
            
            # 如果有上传的图片ID，获取对应的base64数据
            if uploaded_image_id and uploaded_image_id in UPLOADED_IMAGES:
                try:
                    image_base64 = get_image_base64(uploaded_image_id)
                    image_url = None  # 优先使用上传的图片
                except Exception as e:
                    print(f"Warning: Failed to get uploaded image {uploaded_image_id}: {e}")
            
            for attempt in range(max_retries):
                try:
                    # 构建请求体
                    body = {
                        "model": model,
                        "prompt": question,
                        "size": size_value,
                        "sequential_image_generation": payload.sequential or "disabled",
                        "watermark": bool(payload.watermark),
                        "response_format": payload.response_format or "url",
                    }
                    # 提示词优化选项：支持doubao-seedream-4.0系列和doubao-seedream-4.5系列
                    if "doubao-seedream-4.0" in model or "doubao-seedream-4-0" in model or "doubao-seedream-4.5" in model:
                        body["optimize_prompt_options"] = {
                            "mode": payload.optimize_mode or "standard"
                        }
                    
                    # 处理图生图模式：设置 Ark API 所需的 image 参数
                    image_data = None
                    
                    # 优先使用base64图片数据
                    if image_base64:
                        image_data = image_base64
                    # 如果没有base64数据，则从URL下载
                    elif image_url:
                        try:
                            # 从URL下载图片并转换为base64
                            import requests
                            from PIL import Image
                            from io import BytesIO
                            import base64
                            
                            # 下载图片
                            response = requests.get(image_url, timeout=30)
                            response.raise_for_status()
                            
                            # 打开图片
                            image = Image.open(BytesIO(response.content))
                            
                            # 确保图片格式为JPEG
                            if image.mode in ('RGBA', 'LA'):
                                # 创建白色背景
                                background = Image.new('RGB', image.size, (255, 255, 255))
                                # 粘贴图片
                                background.paste(image, mask=image.split()[-1] if image.mode == 'RGBA' else None)
                                image = background
                            
                            # 转换为base64
                            buffer = BytesIO()
                            image.save(buffer, format="JPEG")
                            base64_image = base64.b64encode(buffer.getvalue()).decode('utf-8')
                            
                            # 构建data URL
                            image_data = f"data:image/jpeg;base64,{base64_image}"
                        except Exception as e:
                            print(f"Warning: Failed to process image from URL {image_url}: {e}")
                    
                    # 如果有图片数据，添加到请求体
                    if image_data:
                        body["image"] = image_data
                    
                    if body["sequential_image_generation"] == "auto":
                        body["sequential_image_generation_options"] = {"max_images": payload.max_images or 1}

                    headers = {
                        "Authorization": f"Bearer {api_key}",
                        "Content-Type": "application/json",
                        "User-Agent": "Seedream-Image-Generator/1.0",
                    }

                    url = PICTURE_GENERATION_CONFIG.get("api_endpoint")
                    r = requests.post(url, headers=headers, json=body, timeout=120)  # 增加超时时间到120秒
                    
                    if r.status_code != 200:
                        # 如果是服务器错误（5xx），可以重试
                        if 500 <= r.status_code < 600 and attempt < max_retries - 1:
                            print(f"服务器错误 {r.status_code}，第 {attempt + 1} 次重试")
                            time.sleep(retry_delay * (2 ** attempt))  # 指数退避
                            continue
                        raise RuntimeError(f"图片生成接口状态码: {r.status_code}, 详情: {r.text}")
                    
                    result = r.json()
                    # 确保size字段存在于每个图片对象中
                    for image in result.get("data", []):
                        if "size" not in image:
                            image["size"] = size_value
                    # 添加单张图片的生成耗时
                    result["generation_time"] = round(time.time() - start_time, 2)
                    # 添加输入模式信息
                    result["input_mode"] = payload.input_mode or "text_to_image"
                    return result
                    
                except requests.exceptions.Timeout:
                    if attempt < max_retries - 1:
                        print(f"请求超时，第 {attempt + 1} 次重试")
                        time.sleep(retry_delay * (2 ** attempt))  # 指数退避
                        continue
                    raise RuntimeError("图片生成请求超时（120秒）")
                    
                except requests.exceptions.ConnectionError:
                    if attempt < max_retries - 1:
                        print(f"连接错误，第 {attempt + 1} 次重试")
                        time.sleep(retry_delay * (2 ** attempt))  # 指数退避
                        continue
                    raise RuntimeError("网络连接错误，请检查网络连接")
                    
                except Exception as e:
                    if attempt < max_retries - 1:
                        print(f"未知错误 {type(e).__name__}，第 {attempt + 1} 次重试")
                        time.sleep(retry_delay * (2 ** attempt))  # 指数退避
                        continue
                    raise RuntimeError(f"图片生成失败: {type(e).__name__}: {str(e)}")
            
            # 如果所有重试都失败
            raise RuntimeError("图片生成失败，已达到最大重试次数")

        # 批量并发处理
        from concurrent.futures import ThreadPoolExecutor, as_completed
        results = [None] * len(payload.questions)  # 预分配结果数组，保持顺序
        
        with ThreadPoolExecutor(max_workers=qps) as ex:
            # 提交所有任务
            futures = {ex.submit(single_picture_call, question_data): i for i, question_data in enumerate(payload.questions)}
            
            # 处理完成的任务，保持原始顺序
            for future in as_completed(futures):
                question_index = futures[future]
                question_data = payload.questions[question_index]
                
                # 提取问题文本
                if isinstance(question_data, dict):
                    question_text = question_data.get("question", "")
                else:
                    question_text = question_data
                
                try:
                    # 设置超时时间，避免长时间等待
                    result = future.result(timeout=120)  # 增加超时时间到120秒
                    results[question_index] = {
                        "question": question_text,
                        "result": result,
                        "success": True
                    }
                except Exception as e:
                    results[question_index] = {
                        "question": question_text,
                        "error": str(e),
                        "success": False
                    }
        
        # 过滤掉None值（理论上不应该有，但为了安全）
        results = [r for r in results if r is not None]
        
        # 如果需要下载Excel文件
        if payload.download_excel:
            try:
                # 创建临时目录保存图片和Excel文件
                with tempfile.TemporaryDirectory() as temp_dir:
                    # 生成Excel文件
                    excel_path = create_excel_with_images(results, temp_dir)
                    
                    # 读取Excel文件内容
                    with open(excel_path, "rb") as f:
                        excel_content = f.read()
                    
                    # 返回Excel文件下载
                    return StreamingResponse(
                        BytesIO(excel_content),
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={
                            "Content-Disposition": f"attachment; filename=批量图片生成结果_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
                        }
                    )
            except Exception as e:
                # 如果Excel生成失败，返回原始JSON结果
                return JSONResponse({"error": f"Excel生成失败: {type(e).__name__}: {str(e)}", "results": results}, status_code=500)
        else:
            # 返回JSON结果
            return {"results": results}
        
    except Exception as e:
        return JSONResponse({"error": f"批量图片生成调用失败: {type(e).__name__}: {str(e)}"}, status_code=500)


@app.post("/api/generate_excel")
def generate_excel(payload: ExcelGenPayload):
    """Generate Excel file with images
    
    Args:
        payload: Request data containing image generation results
    
    Returns:
        StreamingResponse: Excel file with images
    """
    try:
        # Validate input data
        if not payload.results or not isinstance(payload.results, list):
            return JSONResponse({"error": "Invalid results data format"}, status_code=400)
        
        # Prepare data format for Excel generation
        excel_data = []
        
        # Process batch generation results
        for i, result in enumerate(payload.results):
            if isinstance(result, dict):
                # Format for single image generation result
                if result.get("data"):
                    excel_data.append({
                        "question": payload.questions[i] if payload.questions and i < len(payload.questions) else f"Image Generation {i+1}",
                        "result": result,
                        "success": True
                    })
                elif result.get("error"):
                    excel_data.append({
                        "question": payload.questions[i] if payload.questions and i < len(payload.questions) else f"Image Generation {i+1}",
                        "error": result.get("error"),
                        "success": False
                    })
            else:
                # Process other result formats
                excel_data.append({
                    "question": payload.questions[i] if payload.questions and i < len(payload.questions) else f"Image Generation {i+1}",
                    "error": "Invalid result data",
                    "success": False
                })
        
        # Create temporary directory to save images and Excel file
        with tempfile.TemporaryDirectory() as temp_dir:
            # Generate Excel file
            excel_path = create_excel_with_images(excel_data, temp_dir)
            
            # Read Excel file content
            with open(excel_path, "rb") as f:
                excel_content = f.read()
            
            # Return Excel file for download
            return StreamingResponse(
                BytesIO(excel_content),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=image_generation_results_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
                }
            )
    except Exception as e:
        # Detailed error handling
        return JSONResponse({"error": f"Excel generation failed: {type(e).__name__}: {str(e)}"}, status_code=500)


@app.post("/api/video_generate")
def video_generate(payload: VideoGenPayload):
    try:
        api_key = payload.api_key or VIDEO_GENERATION_CONFIG.get("api_key")
        if not api_key:
            return JSONResponse({"error": "缺少 API Key"}, status_code=400)

        # 处理并发参数
        run_count = max(1, int(payload.run_count or 1))
        qps = int(payload.qps or VIDEO_GENERATION_CONFIG.get("default_qps", 2))

        def single_video_call():
            """单次视频生成调用"""
            # 默认模型
            model = payload.model or list(VIDEO_GENERATION_CONFIG.get("models", {}).values())[0]
            
            # 确定默认比例
            default_ratio = VIDEO_GENERATION_CONFIG.get("default_ratio", "16:9")
            if payload.input_mode == "图生视频-首帧":
                default_ratio = VIDEO_GENERATION_CONFIG.get("default_i2v_ratio", "adaptive")
            elif payload.input_mode == "图生视频-参考图":
                default_ratio = "16:9"
            
            ratio = payload.ratio or default_ratio
            duration = payload.duration or 5
            resolution = payload.resolution or "720p"

            # 构建请求体
            body = {
                "model": model,
                "content": [],
                "ratio": ratio,
                "duration": duration,
                "resolution": resolution,
            }

            # 添加文本内容
            if payload.input:
                body["content"].append({"type": "text", "text": payload.input})
            
            # 处理图片
            images = []
            if payload.image_urls:
                images = payload.image_urls
            elif payload.image_url:
                images = [payload.image_url]

            if images:
                if payload.input_mode == "图生视频-首尾帧":
                    if len(images) >= 1:
                        body["content"].append({"type": "image_url", "image_url": {"url": images[0]}, "role": "first_frame"})
                    if len(images) >= 2:
                        # Use the last image as the last frame
                        body["content"].append({"type": "image_url", "image_url": {"url": images[-1]}, "role": "last_frame"})
                elif payload.input_mode == "图生视频-参考图":
                    for img_url in images[:4]: # 最多4张
                         body["content"].append({"type": "image_url", "image_url": {"url": img_url}, "role": "reference_image"})
                else: # 默认首帧
                    for img_url in images[:1]:
                        body["content"].append({"type": "image_url", "image_url": {"url": img_url}, "role": "first_frame"})

            # Seedance 1.5 pro 特有参数
            if "1-5-pro" in model:
                if payload.draft is not None:
                    body["draft"] = payload.draft
                if payload.generate_audio is not None:
                    body["generate_audio"] = payload.generate_audio
            
            if payload.return_last_frame:
                body["return_last_frame"] = True

            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
                "User-Agent": "Seedance-Video-Generator/1.0",
            }

            url = VIDEO_GENERATION_CONFIG.get("api_endpoint")
            r = requests.post(url, headers=headers, json=body, timeout=60)
            if r.status_code != 200:
                raise RuntimeError(f"视频生成接口状态码: {r.status_code}, 详情: {r.text}")

            # 获取任务ID
            task_result = r.json()
            task_id = task_result.get("id")
            if not task_id:
                raise RuntimeError(f"视频生成接口未返回任务ID, 详情: {task_result}")

            # 轮询任务状态
            status_url = f"{url}/{task_id}"
            start_time = time.time()
            max_polling_time = 1800  # 最大轮询时间：30分钟
            while time.time() - start_time < max_polling_time:
                status_response = requests.get(status_url, headers=headers, timeout=30)
                if status_response.status_code == 200:
                    status_result = status_response.json()
                    status = status_result.get("status")
                    if status == "succeeded":
                        # 计算耗时
                        total_time = time.time() - start_time
                        # 添加耗时和任务ID到结果
                        status_result["generation_time"] = round(total_time, 2)
                        status_result["task_id"] = task_id
                        return status_result
                    elif status == "failed":
                        raise RuntimeError(f"视频生成失败, 详情: {status_result}")
                else:
                    raise RuntimeError(f"查询任务状态失败: {status_response.status_code}, 详情: {status_response.text}")
                # 间隔2秒轮询一次
                time.sleep(2)

            # 轮询超时
            raise RuntimeError(f"视频生成任务超时, 任务ID: {task_id}")

        # 单次请求直接返回
        if run_count == 1:
            try:
                result = single_video_call()
                return {"results": [result]}
            except Exception as e:
                return JSONResponse({"error": str(e)}, status_code=500)
        
        # 并发请求
        else:
            from concurrent.futures import ThreadPoolExecutor, as_completed
            results = []
            
            with ThreadPoolExecutor(max_workers=qps) as ex:
                futures = [ex.submit(single_video_call) for _ in range(run_count)]
                for future in as_completed(futures):
                    try:
                        results.append(future.result())
                    except Exception as e:
                        results.append({"error": str(e)})
            
            return {"results": results}
    except Exception as e:
        return JSONResponse({"error": f"视频生成调用失败: {type(e).__name__}: {str(e)}"}, status_code=500)


# =============================================================================
# 即梦AI相关功能
# =============================================================================

import hashlib
import hmac

# 即梦AI请求体模型
class JimengAIPayload(BaseModel):
    input: str
    size: int | None = None
    width: int | None = None
    height: int | None = None
    run_count: int | None = 1

# 签名生成辅助函数
def sign(key, msg):
    return hmac.new(key, msg.encode('utf-8'), hashlib.sha256).digest()

def getSignatureKey(key, dateStamp, regionName, serviceName):
    kDate = sign(key.encode('utf-8'), dateStamp)
    kRegion = sign(kDate, regionName)
    kService = sign(kRegion, serviceName)
    kSigning = sign(kService, 'request')
    return kSigning

def formatQuery(parameters):
    request_parameters_init = ''
    for key in sorted(parameters):
        request_parameters_init += key + '=' + parameters[key] + '&'
    request_parameters = request_parameters_init[:-1]
    return request_parameters

# 即梦AI签名和请求函数
def signV4Request(access_key, secret_key, service, req_query, req_body):
    if access_key is None or secret_key is None:
        raise ValueError('No access key is available.')

    t = datetime.utcnow()
    current_date = t.strftime('%Y%m%dT%H%M%SZ')
    datestamp = t.strftime('%Y%m%d')  # Date w/o time, used in credential scope
    canonical_uri = '/'
    canonical_querystring = req_query
    signed_headers = 'content-type;host;x-content-sha256;x-date'
    payload_hash = hashlib.sha256(req_body.encode('utf-8')).hexdigest()
    content_type = 'application/json'
    canonical_headers = 'content-type:' + content_type + '\n' + 'host:' + JIMENG_AI_CONFIG['host'] + '\n' + 'x-content-sha256:' + payload_hash + '\n' + 'x-date:' + current_date + '\n'
    canonical_request = 'POST' + '\n' + canonical_uri + '\n' + canonical_querystring + '\n' + canonical_headers + '\n' + signed_headers + '\n' + payload_hash
    algorithm = 'HMAC-SHA256'
    credential_scope = datestamp + '/' + JIMENG_AI_CONFIG['region'] + '/' + service + '/' + 'request'
    string_to_sign = algorithm + '\n' + current_date + '\n' + credential_scope + '\n' + hashlib.sha256(canonical_request.encode('utf-8')).hexdigest()
    signing_key = getSignatureKey(secret_key, datestamp, JIMENG_AI_CONFIG['region'], service)
    signature = hmac.new(signing_key, (string_to_sign).encode('utf-8'), hashlib.sha256).hexdigest()

    authorization_header = algorithm + ' ' + 'Credential=' + access_key + '/' + credential_scope + ', ' + 'SignedHeaders=' + signed_headers + ', ' + 'Signature=' + signature
    headers = {
        'X-Date': current_date,
        'Authorization': authorization_header,
        'X-Content-Sha256': payload_hash,
        'Content-Type': content_type
    }

    # 发送请求
    request_url = JIMENG_AI_CONFIG['endpoint'] + '?' + canonical_querystring

    try:
        r = requests.post(request_url, headers=headers, data=req_body, timeout=30)
        r.raise_for_status()
    except Exception as err:
        raise RuntimeError(f'请求失败: {str(err)}')
    else:
        # 使用 replace 方法将 \u0026 替换为 &
        resp_str = r.text.replace("\\u0026", "&")
        
        # 解析响应
        task_id = None
        try:
            resp_json = json.loads(resp_str)
            if resp_json.get('code') == 10000 and 'data' in resp_json:
                # 提取task_id
                task_id = resp_json['data'].get('task_id')
        except Exception as e:
            raise RuntimeError(f'解析响应失败: {e}')
        
        return task_id, resp_str

# 查询即梦AI任务结果
def queryTaskResult(access_key, secret_key, service, task_id):
    # 请求Query
    query_params = {
        'Action': 'CVSync2AsyncGetResult',
        'Version': '2022-08-31',
    }
    formatted_query = formatQuery(query_params)

    # 请求Body
    body_params = {
        "req_key": JIMENG_AI_CONFIG['req_key'],
        "task_id": task_id,
        "req_json": "{\"return_url\":true}"
    }
    formatted_body = json.dumps(body_params)
    
    # 调用signV4Request并获取响应
    task_id, resp_str = signV4Request(access_key, secret_key, service, formatted_query, formatted_body)
    return resp_str

# 自动查询即梦AI任务状态
def autoQueryTask(access_key, secret_key, service, task_id, start_time=None):
    max_retries = 60  # 最多查询60次
    retry_interval = 1  # 每1秒查询一次
    
    for _ in range(max_retries):
        # 调用查询接口
        resp_str = queryTaskResult(access_key, secret_key, service, task_id)
        
        # 解析响应状态
        try:
            resp_json = json.loads(resp_str)
            if resp_json.get('code') == 10000 and 'data' in resp_json:
                status = resp_json['data'].get('status')
                
                if status == 'done':
                    # 计算总耗时
                    if start_time:
                        end_time = datetime.now()
                        generation_time = round((end_time - start_time).total_seconds(), 1)
                        # 将耗时信息添加到响应中
                        resp_json['generation_time'] = generation_time
                        return json.dumps(resp_json)
                    return resp_str
                elif status == 'in_queue' or status == 'generating':
                    time.sleep(retry_interval)
                elif status == 'not_found':
                    raise RuntimeError('任务未找到')
                elif status == 'expired':
                    raise RuntimeError('任务已过期')
                else:
                    raise RuntimeError(f'未知状态: {status}')
        except json.JSONDecodeError as e:
            raise RuntimeError(f'解析响应失败: {e}')
        except Exception as e:
            raise e
    
    raise RuntimeError('查询任务超时')

# 即梦AI图片生成API端点
@app.post("/api/jimeng_ai_generate")
def jimeng_ai_generate(payload: JimengAIPayload):
    try:
        # 获取认证信息
        access_key = JIMENG_AI_CONFIG.get('access_key')
        secret_key = JIMENG_AI_CONFIG.get('secret_key')
        
        if not access_key or not secret_key:
            return JSONResponse({"error": "缺少即梦AI的Access Key或Secret Key"}, status_code=400)

        # 处理并发参数
        run_count = max(1, int(payload.run_count or 1))
        qps = JIMENG_AI_CONFIG.get('qps')

        def single_jimeng_call():
            """单次即梦AI调用"""
            # 记录开始时间
            start_time = datetime.now()
            
            # 请求Query
            query_params = {
                'Action': 'CVSync2AsyncSubmitTask',
                'Version': '2022-08-31',
            }
            formatted_query = formatQuery(query_params)

            width = payload.width
            height = payload.height

            if width is not None and height is not None:
                if not (1024 <= width <= 4096):
                    raise RuntimeError('width参数超出范围，必须在1024到4096之间')
                if not (1024 <= height <= 4096):
                    raise RuntimeError('height参数超出范围，必须在1024到4096之间')
                size = width * height
            else:
                size = payload.size or JIMENG_AI_CONFIG.get('default_size', 4194304)

            if not (1048576 <= size <= 16777216):
                raise RuntimeError('size参数超出范围，必须在1024*1024(1048576)到4096*4096(16777216)之间')

            # 请求Body
            req_json = {"return_url": True}
            body_params = {
                "req_key": JIMENG_AI_CONFIG['req_key'],
                "req_json": json.dumps(req_json),
                "prompt": payload.input,
                "size": size
            }
            
            if width is not None and height is not None:
                body_params['width'] = width
                body_params['height'] = height
            formatted_body = json.dumps(body_params)
            
            # 提交任务并获取task_id
            task_id, resp_str = signV4Request(access_key, secret_key, JIMENG_AI_CONFIG['service'], formatted_query, formatted_body)
            
            if not task_id:
                raise RuntimeError('未获取到task_id，任务提交失败')
            
            # 自动查询任务结果（传递开始时间）
            result_str = autoQueryTask(access_key, secret_key, JIMENG_AI_CONFIG['service'], task_id, start_time)
            
            # 解析最终结果
            result_json = json.loads(result_str)
            if result_json.get('code') == 10000 and 'data' in result_json:
                data = result_json['data']
                if data.get('status') == 'done':
                    # 提取图片URL
                    image_urls = data.get('image_urls', [])
                    binary_data = data.get('binary_data_base64', [])
                    
                    result = {
                        "task_id": task_id,
                        "status": data.get('status'),
                        "image_urls": image_urls,
                        "binary_data_base64": binary_data,
                        "message": result_json.get('message', 'Success'),
                        "request_id": result_json.get('request_id'),  # 正确提取request_id
                        "generation_time": result_json.get('generation_time', 0)
                    }
                    return result
                else:
                    raise RuntimeError(f'任务未完成，状态: {data.get("status")}')
            else:
                raise RuntimeError(f'任务执行失败: {result_json.get("message", "未知错误")}')

        # 单次请求直接返回
        if run_count == 1:
            try:
                result = single_jimeng_call()
                return {"results": [result]}
            except Exception as e:
                return JSONResponse({"error": str(e)}, status_code=500)
        
        # 并发请求
        else:
            from concurrent.futures import ThreadPoolExecutor, as_completed
            results = []
            
            with ThreadPoolExecutor(max_workers=qps) as ex:
                futures = [ex.submit(single_jimeng_call) for _ in range(run_count)]
                for future in as_completed(futures):
                    try:
                        results.append(future.result())
                    except Exception as e:
                        results.append({"error": str(e)})
            
            return {"results": results}
    except Exception as e:
        return JSONResponse({"error": f"即梦AI调用失败: {type(e).__name__}: {str(e)}"}, status_code=500)


@app.post("/api/export_results")
def export_results(payload: ExportResultsPayload):
    """导出EP模型推理或联网问答结果为JSON或CSV格式"""
    try:
        def pick_value(obj: dict, keys: list[str], default=""):
            for key in keys:
                if key in obj:
                    val = obj.get(key)
                    if val is None:
                        continue
                    if isinstance(val, str) and val == "":
                        continue
                    return val
            return default

        results = payload.results
        export_format = payload.format.lower()
        result_type = payload.type.lower()
        
        if not results:
            return JSONResponse({"error": "没有可导出的结果"}, status_code=400)
        
        if export_format not in ["json", "csv"]:
            return JSONResponse({"error": "不支持的导出格式，请使用json或csv"}, status_code=400)
        
        if result_type not in ["ep", "network_qa"]:
            return JSONResponse({"error": "不支持的结果类型，请使用ep或network_qa"}, status_code=400)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"doubao_batch_results_{timestamp}"
        
        if export_format == "json":
            content = json.dumps(results, ensure_ascii=False, indent=2)
            media_type = "application/json"
            filename_with_ext = f"{filename}.json"
        else:
            media_type = "text/csv"
            filename_with_ext = f"{filename}.csv"
            
            output = StringIO()
            writer = csv.writer(output)
            
            if result_type == "ep":
                include_reasoning = any(
                    pick_value(r, ["thinking_mode", "thinkingMode", "thinking"], "") not in ["", "disabled"]
                    for r in results
                    if isinstance(r, dict)
                )

                header = [
                    "请求ID",
                    "问题",
                    "回答",
                    "提示词",
                    "总Tokens",
                    "提示词Tokens",
                    "完成Tokens",
                    "TTFT",
                    "总耗时",
                    "模型",
                    "Thinking模式",
                ]
                if include_reasoning:
                    header.append("思考内容")
                header.extend(["Temperature", "Top P", "Max Tokens"])
                writer.writerow(header)
                for result in results:
                    usage = result.get("usage", {})
                    req_id = pick_value(result, ["id", "request_id", "requestId"], "")
                    model = pick_value(result, ["model", "model_id", "modelId"], "")
                    thinking_mode = pick_value(result, ["thinking_mode", "thinkingMode", "thinking"], "")
                    reasoning = pick_value(result, ["reasoning", "reasoning_content", "reasoningContent"], "")
                    temperature = pick_value(result, ["temperature"], "")
                    top_p = pick_value(result, ["top_p", "topP"], "")
                    max_tokens = pick_value(result, ["max_tokens", "maxTokens"], "")

                    row = [
                        req_id,
                        result.get("question", ""),
                        result.get("content", ""),
                        result.get("system_prompt", ""),
                        usage.get("total_tokens", ""),
                        usage.get("prompt_tokens", ""),
                        usage.get("completion_tokens", ""),
                        result.get("first_token_time", ""),
                        result.get("total_time", ""),
                        model,
                        thinking_mode,
                    ]
                    if include_reasoning:
                        row.append(reasoning if thinking_mode not in ["", "disabled"] else "")
                    row.extend([temperature, top_p, max_tokens])

                    writer.writerow(row)
            else:
                writer.writerow(["请求ID", "问题", "回答", "提示词", "总Tokens", "提示词Tokens", "完成Tokens", "TTFT", "总耗时", "Thinking模式"])
                for result in results:
                    usage = result.get("usage", {})
                    req_id = pick_value(result, ["id", "request_id", "requestId"], "")
                    writer.writerow([
                        req_id,
                        result.get("question", ""),
                        result.get("content", ""),
                        result.get("system_prompt", ""),
                        usage.get("total_tokens", ""),
                        usage.get("prompt_tokens", ""),
                        usage.get("completion_tokens", ""),
                        result.get("first_token_time", ""),
                        result.get("total_time", ""),
                        result.get("thinking", "")
                    ])
            
            output.seek(0)
            content = output.getvalue()
        
        return StreamingResponse(
            iter([content.encode('utf-8')]),
            media_type=media_type,
            headers={
                "Content-Disposition": f'attachment; filename="{filename_with_ext}"'
            }
        )
    except Exception as e:
        return JSONResponse({"error": f"导出失败: {type(e).__name__}: {str(e)}"}, status_code=500)
